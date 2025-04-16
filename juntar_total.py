import os
from openpyxl import load_workbook

import pandas as pd
# Script solicitado para merge dos arquivos .dat

def read_dat_file(filepath, offset=4):
    """
    Lê um arquivo .dat e retorna o cabeçalho e os dados.
    
    :param filepath: Caminho para o arquivo .dat
    :param offset: Número de linhas do cabeçalho
    :return: Tuple contendo o cabeçalho e os dados
    """
    with open(filepath, 'r', encoding='utf-8') as file:
        lines = file.readlines()
    header = lines[:offset]
    data = lines[offset:]
    return header, data

def list_dat_files(directory):
    """
    Lista todos os arquivos .dat e .backup em um diretório.
    
    :param directory: Caminho para o diretório
    :return: Lista de arquivos .dat e .backup
    """
    return [f for f in os.listdir(directory) if f.endswith('.dat') or f.endswith('.backup')]

def merge_dat_files(directory, output_file):
    """
    Mescla todos os arquivos .dat em um diretório em um único arquivo.
    
    :param directory: Caminho para o diretório contendo os arquivos .dat
    :param output_file: Caminho para o arquivo de saída
    """
    dat_files = list_dat_files(directory)

    if not dat_files:
        print("Nenhum arquivo .dat encontrado no diretório.")
        return

    all_data = []
    common_header = None

    for i, dat_file in enumerate(dat_files):
        filepath = os.path.join(directory, dat_file)
        header, data = read_dat_file(filepath)

        # Para o primeiro arquivo, salvar o cabeçalho e os dados
        if i == 0:
            common_header = header
            all_data.extend(header + data)
        else:
            all_data.extend(data)
            # Para arquivos subsequentes, verificar se o cabeçalho é o mesmo
            if header != common_header:
                print(f"Erro: Cabeçalho do arquivo {dat_file} é diferente do cabeçalho comum.")
                # print(f"Problema com o arquivo: {dat_file}")
                # print("HEADER: ", header)
                # print("COMMOM HEADER: ", common_header)
                continue

    # Escrever todos os dados no arquivo de saída
    with open(output_file, 'w', encoding='utf-8') as file:
        file.writelines(all_data)

    print(f"Arquivo '{output_file}' criado com sucesso.")

    columns = all_data[:4]  # Divide a string para obter uma lista de colunas
    data = [line.strip().split(",") for line in all_data[4:]]  # Divide cada linha dos dados




    # Crie o DataFrame
    df = pd.DataFrame(data)

    # Remove aspas duplas da primeira coluna (data)
    df.iloc[:, 0] = df.iloc[:, 0].str.strip('"')

    # Ordena o DataFrame pela primeira coluna (data)
    df.sort_values(by=df.columns[0], inplace=True)

    # Converte a primeira coluna para o formato de data e depois para o formato desejado
    df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], format="%Y-%m-%d %H:%M:%S").dt.strftime("%d/%m/%Y %H:%M:%S")

    # Formata as colunas numéricas para duas casas decimais e substitui '.' por ','
    for col in df.columns[1:]:
        df[col] = df[col].apply(
            lambda x: f"{float(x):.2f}".replace('.', ',') if x.replace('.', '', 1).isdigit() else x
        )

    # Cria o diretório 'consolidado' se ele não existir
    if not os.path.exists('consolidado'):
        os.makedirs('consolidado')

    # Salva o DataFrame em um arquivo Excel
    #print(output_file)
    estacao = str(output_file).split("-")[-2]
    df.to_excel(f"consolidado/consolidado_{estacao}.xlsx", index=False, header=False)

    # Abre o arquivo Excel e insere o cabeçalho
    wb = load_workbook(f"consolidado/consolidado_{estacao}.xlsx")
    ws = wb.active

    for i, linha in enumerate(columns, 1):
        valores = linha.split(',')
        ws.insert_rows(i)  # Insere uma nova linha no topo
        for j, valor in enumerate(valores, 1):
            valor = valor.replace('"', '')
            ws.cell(row=i, column=j, value=valor)

    # Salva o arquivo com as novas linhas de cabeçalho
    wb.save(f"consolidado/consolidado_{estacao}.xlsx")

if __name__ == "__main__":
    # Diretório contendo arquivos
    directory = 'juntar/RNES03-2024-12'
    
    # Arquivo de saída
    output_file = 'RNES03-2024-12.dat'

    # Mesclar arquivos
    merge_dat_files(directory, output_file)
